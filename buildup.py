import pandas as pd
import openpyxl
import datetime



def build_corpus()->dict:
    """build the raw corpus from the excel sheet in the `onstart` folder"""
    wb = openpyxl.load_workbook('data/onstart/WSM_FinancialStmt_evolution_copy.xlsx', data_only=True)
    corpus = {}
    all_months = wb.sheetnames
    for month in all_months:    
        sheet = wb[month]
        corpus[month] = {}
        create_category = True
        for cell in sheet['A']:
            if not cell.value:
                continue 

            b_cell = sheet.cell(row=cell.row, column=2)

            if b_cell.value is not None:
                value = b_cell.value
                corpus[month][current_category][cell.value] = value
                if not create_category:
                    create_category = True
            else:
                if create_category:
                    if str(cell.value).isupper():
                        current_category = cell.value
                        corpus[month][current_category] = {}
                        create_category = False

    return corpus

def get_corpus_dates(corpus:dict)-> dict:
    """Add dates to the corpus upper level keys"""
    corpus_with_dates = {}

    for month in corpus:
        month_str = datetime.datetime.strptime(month.split('-')[1], "%B").strftime("%b")
        month_year = month_str+'-2022'
        datetime.datetime.strptime(month_year, '%b-%Y')
        corpus_with_dates[month_year] = corpus[month]

    return corpus_with_dates


def get_options(corpus:dict)-> tuple[list, dict]:
    """retrieve a list of upper level categories and a dictionary with options beneath each categories"""
    categories = []

    for key in corpus:
        cats = (list(corpus[key].keys()))
        categories += [x for x in cats if x not in categories]

    cat_options = {}

    for cat in categories:
        cat_options[cat] = []
        for date in corpus.keys():
            items = list(corpus[date][cat].keys())
            cat_options[cat] += [x for x in items if x not in cat_options[cat]]

    return categories, cat_options

def main():
    """Get the main corpus with datetime keys"""
    corpus = build_corpus()
    corpus = get_corpus_dates(corpus)

    return corpus


if __name__ =="__main__":
    print('This is a library of functions for the app.py file, not a script. Please import the functions to use.')